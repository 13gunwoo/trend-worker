# exporters/munpia_to_excel.py

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CELL_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _write_row(ws, values, row):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT
        cell.border = BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_best_sheet(wb, sheet_name, items):
    ws = wb.create_sheet(title=sheet_name)
    headers = ["순위", "제목", "작가", "장르", "URL", "소개글"]
    _write_header(ws, headers)
    for i, item in enumerate(items, 2):
        _write_row(ws, [
            item.get("rank", ""),
            item.get("title", ""),
            item.get("author", ""),
            item.get("genre", ""),
            item.get("url", ""),
            item.get("description", "") or "",
        ], i)
    _auto_width(ws)


def export(data, output_dir):
    today = date.today().strftime("%Y%m%d")
    platform = data.get("플랫폼", "unknown")
    filename = platform + "_" + today + ".xlsx"
    filepath = os.path.join(output_dir, filename)

    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    category_map = {
        "무료투데이베스트": "무료투데이",
        "유료투데이베스트": "유료투데이",
        "유료신규베스트"  : "유료신규",
        "유료인기급상승"  : "유료급상승",
    }

    for key, sheet_name in category_map.items():
        if key in data:
            _write_best_sheet(wb, sheet_name, data[key])

    wb.save(filepath)
    return filepath
