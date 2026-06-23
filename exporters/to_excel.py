# exporters/to_excel.py

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CELL_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _write_row(ws, values, row):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT
        cell.border = BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_ranking_sheet(wb, sheet_name, items):
    ws = wb.create_sheet(title=sheet_name)
    headers = ["순위", "변동", "작품명", "작가", "장르", "조회수", "연재상태", "연령등급", "series_id"]
    _write_header(ws, headers)
    for i, item in enumerate(items, 2):
        _write_row(ws, [
            item.get("순위", ""),
            item.get("변동", ""),
            item.get("작품명", ""),
            item.get("작가", ""),
            item.get("장르", ""),
            item.get("조회수", ""),
            item.get("연재상태", ""),
            item.get("연령등급", ""),
            item.get("series_id", ""),
        ], i)
    _auto_width(ws)
    return ws


def export(data, output_dir):
    today = date.today().strftime("%Y%m%d")
    platform = data.get("플랫폼", "unknown")
    filename = platform + "_" + today + ".xlsx"
    filepath = os.path.join(output_dir, filename)

    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # 실시간 랭킹
    if "실시간랭킹" in data:
        _write_ranking_sheet(wb, "실시간랭킹", data["실시간랭킹"])

    # 장르별 랭킹
    if "장르별랭킹" in data:
        for genre_name, items in data["장르별랭킹"].items():
            _write_ranking_sheet(wb, genre_name, items)

    # 신작
    if "신작" in data:
        ws = wb.create_sheet(title="신작")
        headers = ["날짜", "작품명", "장르", "조회수", "연령등급", "series_id"]
        _write_header(ws, headers)
        row = 2
        for day_group in data["신작"]:
            date_label = day_group.get("날짜", "")
            for item in day_group.get("작품목록", []):
                _write_row(ws, [
                    date_label,
                    item.get("작품명", ""),
                    item.get("장르", ""),
                    item.get("조회수", ""),
                    item.get("연령등급", ""),
                    item.get("series_id", ""),
                ], row)
                row += 1
        _auto_width(ws)

    # 작품 상세
    if "작품상세" in data:
        ws = wb.create_sheet(title="작품상세")
        headers = ["series_id", "작품명", "작가", "장르", "연령등급", "연재상태", "연재주기",
                   "조회수", "댓글수", "별점", "무료회차", "결제방식", "키워드"]
        _write_header(ws, headers)
        for i, item in enumerate(data["작품상세"], 2):
            _write_row(ws, [
                item.get("series_id", ""),
                item.get("작품명", ""),
                item.get("작가", ""),
                item.get("장르", ""),
                item.get("연령등급", ""),
                item.get("연재상태", ""),
                item.get("연재주기", ""),
                item.get("조회수", 0),
                item.get("댓글수", 0),
                item.get("별점", ""),
                item.get("무료회차", 0),
                item.get("결제방식", ""),
                item.get("키워드", ""),
            ], i)
        _auto_width(ws)

    wb.save(filepath)
    return filepath
