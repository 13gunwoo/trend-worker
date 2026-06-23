# exporters/naver_to_excel.py

import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CELL_FONT   = Font(name="Arial", size=10)
THIN        = Side(style="thin", color="CCCCCC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _write_row(ws, values, row):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = CELL_FONT
        cell.alignment = LEFT
        cell.border = BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_ranking_sheet(wb, sheet_name, items):
    ws = wb.create_sheet(title=sheet_name[:31])  # 시트명 31자 제한
    headers = ["순위", "제목", "작가", "URL", "순위상승", "완결여부", "성인여부"]
    _write_header(ws, headers)
    for i, item in enumerate(items, 2):
        _write_row(ws, [
            item.get("rank", ""),
            item.get("title", ""),
            item.get("author", ""),
            item.get("url", ""),
            "Y" if item.get("isUp") else "",
            "완결" if item.get("isFinished") else "연재중",
            "Y" if item.get("isAdult") else "",
        ], i)
    _auto_width(ws)


def _write_list_sheet(wb, sheet_name, items):
    ws = wb.create_sheet(title=sheet_name[:31])
    headers = ["순위", "제목", "작가", "별점", "URL", "신작여부", "완결여부", "성인여부"]
    _write_header(ws, headers)
    for i, item in enumerate(items, 2):
        _write_row(ws, [
            item.get("rank", ""),
            item.get("title", ""),
            item.get("author", ""),
            item.get("starScore", ""),
            item.get("url", ""),
            "Y" if item.get("isNew") else "",
            "완결" if item.get("isFinished") else "연재중",
            "Y" if item.get("isAdult") else "",
        ], i)
    _auto_width(ws)


def export(data, output_dir):
    today = date.today().strftime("%Y%m%d")
    platform = data.get("플랫폼", "unknown")
    filename = platform + "_" + today + ".xlsx"
    filepath = os.path.join(output_dir, filename)

    os.makedirs(output_dir, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    # 랭킹 (전체/남성/여성)
    if "랭킹" in data:
        for label, items in data["랭킹"].items():
            if items:
                _write_ranking_sheet(wb, "랭킹_" + label, items)

    # 장르별 랭킹
    if "장르별랭킹" in data:
        for genre, items in data["장르별랭킹"].items():
            if items:
                _write_list_sheet(wb, "장르_" + genre, items)

    # 요일별 랭킹
    if "요일별랭킹" in data:
        for day, items in data["요일별랭킹"].items():
            if items:
                _write_list_sheet(wb, "요일_" + day, items)

    # 검색결과
    if "검색결과" in data:
        query = data["검색결과"].get("검색어", "")
        items = data["검색결과"].get("목록", [])
        if items:
            ws = wb.create_sheet(title=("검색_" + query)[:31])
            headers = ["순위", "제목", "작가", "장르", "URL", "완결여부", "성인여부"]
            _write_header(ws, headers)
            for i, item in enumerate(items, 2):
                genres = ", ".join(item.get("genres", []))
                _write_row(ws, [
                    item.get("rank", ""),
                    item.get("title", ""),
                    item.get("author", ""),
                    genres,
                    item.get("url", ""),
                    "완결" if item.get("isFinished") else "연재중",
                    "Y" if item.get("isAdult") else "",
                ], i)
            _auto_width(ws)

    # 트렌드 분석
    if "트렌드분석" in data:
        # 트렌드 태그
        tags = data["트렌드분석"].get("트렌드태그", [])
        if tags:
            ws = wb.create_sheet(title="트렌드태그")
            _write_header(ws, ["태그", "빈도수"])
            for i, t in enumerate(tags, 2):
                _write_row(ws, [t.get("tag", ""), t.get("count", 0)], i)
            _auto_width(ws)

        # 인기작
        popular = data["트렌드분석"].get("인기작", [])
        if popular:
            ws = wb.create_sheet(title="트렌드인기작")
            _write_header(ws, ["제목", "구독자수", "태그", "요일랭킹", "URL"])
            for i, item in enumerate(popular, 2):
                _write_row(ws, [
                    item.get("title", ""),
                    item.get("favoriteCount", 0),
                    ", ".join(item.get("tags", [])),
                    ", ".join(item.get("dayRanks", [])),
                    item.get("url", ""),
                ], i)
            _auto_width(ws)

    wb.save(filepath)
    return filepath
